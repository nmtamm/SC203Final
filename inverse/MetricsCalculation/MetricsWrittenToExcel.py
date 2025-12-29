import os
import json
from openpyxl import Workbook, load_workbook

# Define the root directory
root_dir = "path to your directory containing JSON files"
excel_file_path = os.path.join(root_dir, "MetricsSummary.xlsx")

# Create Excel file if it doesn't exist
if not os.path.exists(excel_file_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Metrics Summary"
    sheet.append(["Filename", "IoU", "F1", "ROUGE", "BLEU"])
    workbook.save(excel_file_path)

# Load workbook once
workbook = load_workbook(excel_file_path)
sheet = workbook.active

for filename in os.listdir(root_dir):
    if filename.endswith(".json"):
        print(f"Processing file: {filename}")
        new_filename = filename[:-5].capitalize()  # Capitalize first character

        file_path = os.path.join(root_dir, filename)
        with open(file_path, "r", encoding="utf-8") as f:
            notebook = json.load(f)
        metrics = notebook.get("average_metrics", {})
        iou = metrics.get("Average IoU", 0)
        f1 = metrics.get("Average F1", 0)
        rouge = metrics.get("Average ROUGE-L", 0)
        bleu = metrics.get("Average SacreBLEU", 0)

        # Find and update the row with the matching filename (capitalize first character)
        for row in sheet.iter_rows(min_row=2):
            excel_name = str(row[0].value).capitalize()
            if excel_name == new_filename:
                row[1].value = iou
                row[2].value = f1
                row[3].value = bleu
                row[4].value = rouge
                break  # Stop after updating the matching row

workbook.save(excel_file_path)
